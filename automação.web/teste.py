from openpyxl import load_workbook

planilha_fornecedores = load_workbook('./fornecedores.xlsx')
pagina_fornecedores = planilha_fornecedores['Sheet1']

for linha in pagina_fornecedores.iter_rows( values_only= True):
    print(linha)